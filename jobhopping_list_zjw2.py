import pandas as pd
from pandas import Series, DataFrame
import openpyxl

wb = openpyxl.load_workbook(filename=r'E:\python宇宙\python_examples\zjw_linkedin\20201011_test\select\jh-test2.xlsx')
openpyxl.sheet_ranges = wb['Sheet1']
ws = wb.active
# 调用python的excel处理办法，便于每个数据的填写

org_data = pd.read_excel(r'E:\python宇宙\python_examples\zjw_linkedin\20201011_test\select\s-droplist.xlsx')
id_data = list(org_data["id"])
# 把原始数据的id变成list 格式，便于数据处理

frame0 = pd.DataFrame(org_data)
two = frame0[["company_6",'job_6','duration_6','description_6']]
# 调用多个行的方法

obj0 = frame0["company_6"]
isnan = pd.isnull(obj0)
# 对NaN的部分进行Boolean判断，是否为False

time_1 = frame0['duration_1'][0][:8]
# 对数据进行切片处理，得到我们需要的跳槽时间数据
c=0
for m in frame0.index:
    f = dict(frame0.loc[m])
    # print(str((f['company_6']))) 打印转换成字符串的数值
    if str(f['duration_7']) != 'nan':
        n = int(f['id']-1)
        frame0 = frame0.drop([n])
        # print(m,n)
        c+=1
# if条件句，判断是删除我们不需要的行数据duration_7不为空的话就删掉
idx = 0
for i in frame0.index[1583:]:
    f = dict(frame0.loc[i])
    # print(f)
    # 把表格的每一行用index进行读取，转化成‘字典’便于操作
    j = 0
    while j<5:
        j+=1
        clast_x = 'company_' + str(j+1)
        cnext_x = 'company_' + str(j)
        jlast_x = 'job_' + str(j+1)
        jnext_x = 'job_' + str(j)
        d_x = 'duration_' + str(j)
        deslast_x = 'description_' + str(j+1)
        desnext_x = 'description_' + str(j)
        # 创造一系列代表title的变量值
        idn = str(idx+2) # 顺序变量，从1开始，根据idx值变化，便于竖向索引
        ws['A'+idn] = idx
        ws['B'+idn] = str(f[d_x])[:8]
        ws['C'+idn] = str(f['id'])
        ws['D'+idn] = str(f['name'])
        ws['E'+idn] = str(f[clast_x])
        ws['F'+idn] = str(f[cnext_x])
        ws['G'+idn] = str(f[jlast_x])
        ws['H'+idn] = str(f[jnext_x])
        ws['I'+idn] = str(f[deslast_x])
        ws['J'+idn] = str(f[desnext_x])
        idx += 1
wb.save(r'E:\python宇宙\python_examples\zjw_linkedin\20201011_test\select\jh-test2.xlsx')
